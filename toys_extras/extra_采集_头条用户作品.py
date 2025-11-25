import openpyxl
from toys_extras.base_web import BaseWeb
from playwright.sync_api import Page
from toys_logger import logger
from datetime import datetime, timedelta
import os

__version__ = '1.0.3'


class Toy(BaseWeb):

    def __init__(self, page: Page):
        super().__init__(page)
        self.result_table_view: list = [['文章标题', "类型", "链接", "热度", "发布时间", "作者"]]

    def play(self):
        作者主页地址 = self.config.get("扩展", "作者主页地址")
        采集类别 = self.config.get("扩展", "采集类别")
        最低阅读量 = self.config.get("扩展", "最低阅读量")
        发布日期 = self.config.get("扩展", "发布日期")
        发布时间起始 = self.config.get("扩展", "发布时间起始")
        发布时间截止 = self.config.get("扩展", "发布时间截止")
        存储目录 = self.config.get("扩展", "存储目录")
        excel作者主页地址列标题名 = self.config.get("扩展", "excel作者主页地址列标题名")
        if not 作者主页地址 and not self.files:
            return
        urls = set()
        if 作者主页地址:
            urls.add(作者主页地址)
        for file in self.files:
            if not file.endswith(('.xlsx', 'txt')):
                continue
            if file.endswith('.txt'):
                with open(file, 'r', encoding='utf-8') as f:
                    lines = f.read().splitlines()
                for line in lines:
                    urls.add(line.strip())
                continue
            else:
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                headers = next(sheet.iter_rows(values_only=True))
                try:
                    url_index = headers.index(excel作者主页地址列标题名)
                except ValueError:
                    logger.error(f"excel文件{file}中没有找到{excel作者主页地址列标题名}列")
                    continue
                for row in sheet.iter_rows(min_row=2):
                    url_cell = row[url_index]
                    if url_cell.hyperlink:
                        url = url_cell.hyperlink.target
                    else:
                        url = url_cell.value
                    urls.add(url.strip())
        if not 发布日期:
            logger.info("发布日期未设置，不进行采集")
            return
        publish_date = datetime.now() - timedelta(days=int(发布日期))
        if 发布时间起始:
            publish_time_start = datetime.strptime(发布时间起始, '%H:%M:%S')
            publish_time_start = publish_date.replace(hour=publish_time_start.hour, minute=publish_time_start.minute, second=publish_time_start.second, microsecond=0)
        else:
            publish_time_start = publish_date.replace(hour=0, minute=0, second=0, microsecond=0)
        if 发布时间截止:
            publish_time_end = datetime.strptime(发布时间截止, '%H:%M:%S')
            publish_time_end = publish_date.replace(hour=publish_time_end.hour, minute=publish_time_end.minute, second=publish_time_end.second, microsecond=0)
        else:
            publish_time_end = datetime.now().replace(microsecond=0)
        collect_articles = []
        match 采集类别:
            case "文章":
                url_patten = "/list/user/feed?category=pc_profile_article"
            case "视频":
                url_patten = "/list/user/feed?category=pc_profile_video"
            case "微头条":
                url_patten = "/list/user/feed?category=pc_profile_ugc"
            case _:
                url_patten = "/list/user/feed?category=profile_all"
        for url in urls:
            try:
                with self.page.expect_response(lambda response: url_patten in response.url) as response_info:
                    self.page.goto(url)
                    if self.page.title() == "404错误页":
                        self.result_table_view.append([f"{url} 不存在", "", "", "", "", ""])
                        continue
                    type_locator = self.page.get_by_role("tab", name=采集类别)
                    if type_locator.get_attribute("aria-selected") != "true":
                        type_locator.click()
                response = response_info.value
                need_more_data = True
                while need_more_data:
                    articles = response.json()["data"]
                    for article in articles:
                        article_type = article["log_pb"].get("article_type", "")
                        publish_time = article["publish_time"]
                        publish_time = datetime.fromtimestamp(publish_time)
                        if publish_time < publish_time_start:
                            if article is articles[-1]:
                                need_more_data = False
                            continue
                        if publish_time > publish_time_end or publish_time < publish_time_start:
                            continue
                        read_count = article["itemCell"]["itemCounter"]["readCount"]
                        if read_count < int(最低阅读量):
                            continue
                        if article_type == "weitoutiao":
                            title = article['share']["share_title"]
                            link = article["share_url"]
                            author = article["user"]["screen_name"]
                            type_ = "微头条"
                        elif "video_source" in article:
                            title = article['share']["share_title"]
                            link = article["share_url"]
                            author = article["user_info"]["name"]
                            type_ = "视频"
                        else:
                            title = article["title"]
                            link = article["url"]
                            author = article["user_info"]["name"]
                            type_ = "文章"
                        self.result_table_view.append([title, type_, link, read_count, publish_time, author])
                        collect_articles.append([title, type_, link, read_count, publish_time, author])
                    if need_more_data and response.json()["has_more"]:
                        with self.page.expect_response(lambda response: url_patten in response.url) as response_info:
                            self.page.evaluate("window.scrollTo(0, document.body.scrollHeight);")
                            self.page.wait_for_timeout(1000)
                        response = response_info.value
                    else:
                        need_more_data = False
            except Exception as e:
                logger.error(f"打开{url}失败: {e}")
                continue
        self.page.close()
        if not collect_articles:
            logger.info("没有找到符合条件的文章")
            return
        if not os.path.exists(存储目录):
            os.makedirs(存储目录)
        filename = f"{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}_文章列表.xlsx"
        filepath = os.path.join(存储目录, filename)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(self.result_table_view[0])
        for row in collect_articles:
            sheet.append(row)
        workbook.save(filepath)


